import streamlit as st
import openpyxl
import os
import tempfile
from io import BytesIO

# ---------------------------------------------------------------------------
# Expected Headers
# ---------------------------------------------------------------------------

EXPECTED_HEADERS = [
    "#", "Name", "MOPID", "Date", "CO_Ent_Nbr", "CO_Name",
    "Ultimate_Parent_EntNbr", "Ultimate_Parent_Name",
    "Immediate_Parent_EntNbr", "Immediate_Parent_Name",
    "Before_URL", "Company_Status", "Company_Remarks",
    "Additional Comments", "Reference_URL", "Year/Date of Change",
    "Finance Remarks", "AR Link",
    "Before_Address1", "After_Address1", "ADDRESS1ADMV",
    "Before_Address2", "After_Address2", "ADDRESS2ADMV",
    "Before_Address3", "After_Address3", "ADDRESS3ADMV",
    "Before_Postal_Code1", "After_Postal_Code1", "POSTAL_CODE1ADMV",
    "Before_City", "After_City", "CITYADMV",
    "Before_State/Province", "After_State/Province", "State/ProvinceADMV",
    "Before_Zip", "After_Zip", "ZIPADMV",
    "Before_Postal_Code2", "After_Postal_Code2", "POSTAL_CODE2ADMV",
    "Before_Country", "After_Country", "COUNTRYADMV",
    "Before_Postal_Code3", "After_Postal_Code3", "POSTAL_CODE3ADMV",
    "Before_Mailing_Address1", "After_Mailing_Address1", "MAILING_ADDRESS1ADMV",
    "Before_Mailing_Address2", "After_Mailing_Address2", "MAILING_ADDRESS2ADMV",
    "Before_Mailing_Address3", "After_Mailing_Address3", "MAILING_ADDRESS3ADMV",
    "Before_Mailing_PostalCode1", "After_Mailing_PostalCode1", "MAILING_POSTALCODE1ADMV",
    "Before_Mailing_City", "After_Mailing_City", "MAILING_CITYADMV",
    "Before_Mailing_State", "After_Mailing_State", "MAILING_STATEADMV",
    "Before_Mailing_Zip", "After_Mailing_Zip", "MAILING_ZIPADMV",
    "Before_Mailing_Postal_Code2", "After_Mailing_Postal_Code2", "MAILING_POSTAL_CODE2ADMV",
    "Before_Mailing_Country", "After_Mailing_Country", "MAILING_COUNTRYADMV",
    "Before_Mailing_Postal_Code3", "After_Mailing_Postal_Code3", "MAILING_POSTAL_CODE3ADMV",
    "Before_Phone", "After_Phone", "PHONEADMV",
    "Before_Fax", "After_Fax", "FAXADMV",
    "Before_Toll_Free", "After_Toll_Free", "TOLL_FREEADMV",
    "Before_Email", "After_Email", "EMAILADMV",
    "Before_URL", "After_URL", "URLADMV",
    "Before_State_Incorp", "After_State_Incorp", "STATE_INCORPADMV",
    "Before_Year_Founded", "After_Year_Founded", "YEAR_FOUNDEDADMV",
    "Before_Nbr_Employees", "After_Nbr_Employees", "NBR_EMPLOYEESADMV",
    "Before_Business_Descrip", "After_Business_Descrip", "BUSINESS_DESCRIPADMV",
    "Before_Primary_NAICS_Code1", "After_Primary_NAICS_Code1", "PRIMARY_NAICS_CODE1ADMV",
    "Source_URL", "Year_Found_URL", "Emp_Founded Year",
    "State Inc_Url", "Annual Report Status", "Annual Report URL",
]

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

SKIP_ADMV_COLS = {
    "Before_Phone", "After_Phone", "PHONEADMV",
    "Before_Fax", "After_Fax", "FAXADMV",
    "Before_Toll_Free", "After_Toll_Free", "TOLL_FREEADMV",
}

EMPLOYEE_ADMV      = "NBR_EMPLOYEESADMV"
STATUS_NAME_CHANGE = "Problem Type-Company name change"
STATUS_SKIP        = [
    "Problem Type-Company out of business",
    "Non Valid",
    "Problem Type-Merger & Acquisition",
]

# ---------------------------------------------------------------------------
# Core Logic
# ---------------------------------------------------------------------------

def is_empty(val):
    if val is None:
        return True
    s = str(val).strip()
    return s == "" or s.lower() == "nan"

def get_admv(before, after, is_employee=False):
    b_empty = is_empty(before)
    a_empty = is_empty(after)
    if not b_empty and not a_empty:
        return "V" if str(before).strip() == str(after).strip() else "M"
    elif b_empty and not a_empty:
        return "A"
    elif not b_empty and a_empty:
        return "D" if is_employee else "X"
    return None

def should_proceed(status, additional_comments):
    if is_empty(status):
        return True
    status_str = str(status).strip()
    if status_str == "Valid":
        return True
    for skip in STATUS_SKIP:
        if status_str == skip:
            return False
    if status_str == STATUS_NAME_CHANGE:
        comments      = str(additional_comments).strip() if not is_empty(additional_comments) else ""
        has_valid     = "Valid"     in comments
        has_non_valid = "Non Valid" in comments
        if has_valid and has_non_valid:
            return None
        elif has_non_valid:
            return False
        elif has_valid:
            return True
        return None
    return None

def validate_headers(actual_headers):
    actual_set   = set(str(h).strip() for h in actual_headers if h is not None)
    expected_set = set(EXPECTED_HEADERS)
    missing      = sorted(expected_set - actual_set)
    extra        = sorted(actual_set   - expected_set)
    return (len(missing) == 0), missing, extra

def find_admv_groups(header_row):
    col_map = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}
    groups  = []
    for col_name, admv_idx in col_map.items():
        if not col_name.endswith("ADMV") or col_name in SKIP_ADMV_COLS:
            continue
        base       = col_name[:-4]
        before_idx = None
        after_idx  = None
        for c, i in col_map.items():
            if c.lower() == ("before_" + base).lower():
                before_idx = i
            if c.lower() == ("after_"  + base).lower():
                after_idx  = i
        if before_idx is None or after_idx is None:
            continue
        if header_row[before_idx] in SKIP_ADMV_COLS or header_row[after_idx] in SKIP_ADMV_COLS:
            continue
        groups.append((before_idx, after_idx, admv_idx, col_name == EMPLOYEE_ADMV))
    return groups

def process_excel(file_bytes):
    logs    = []
    summary = {}

    tmp_in = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_in.write(file_bytes)
    tmp_in.close()

    wb = openpyxl.load_workbook(tmp_in.name)
    ws = wb.active
    os.unlink(tmp_in.name)

    header_row             = [cell.value for cell in ws[1]]
    is_valid, missing, extra = validate_headers(header_row)

    if not is_valid:
        return False, None, logs, {"missing": missing, "extra": extra}

    col_map      = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}
    status_idx   = col_map.get("Company_Status")
    comments_idx = col_map.get("Additional Comments")
    admv_groups  = find_admv_groups(header_row)

    total_rows  = 0
    skipped     = 0
    left_as_is  = 0
    changes     = 0
    d_protected = 0
    admv_counts = {"V": 0, "M": 0, "A": 0, "X": 0, "D": 0}

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        total_rows += 1
        status   = row[status_idx].value   if status_idx   is not None else None
        comments = row[comments_idx].value if comments_idx is not None else None
        proceed  = should_proceed(status, comments)

        if proceed is False:
            skipped += 1
            logs.append(("skip",  row_num, str(status) if status else "", 0))
            continue
        if proceed is None:
            left_as_is += 1
            logs.append(("leave", row_num, str(status) if status else "", 0))
            continue

        row_changes = 0
        for before_idx, after_idx, admv_idx, is_employee in admv_groups:
            admv_cell    = row[admv_idx]
            current_admv = admv_cell.value
            if not is_empty(current_admv) and str(current_admv).strip().upper() == "D":
                d_protected += 1
                continue
            new_admv = get_admv(row[before_idx].value, row[after_idx].value, is_employee)
            if new_admv is not None:
                admv_cell.value = new_admv
                admv_counts[new_admv] += 1
                row_changes += 1
                changes     += 1

        logs.append(("ok", row_num, str(status) if status else "", row_changes))

    out_buffer = BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)

    summary = {
        "total_rows" : total_rows,
        "skipped"    : skipped,
        "left_as_is" : left_as_is,
        "processed"  : total_rows - skipped - left_as_is,
        "changes"    : changes,
        "d_protected": d_protected,
        "admv_counts": admv_counts,
        "extra"      : extra,
        "missing"    : [],
    }

    return True, out_buffer.getvalue(), logs, summary

# ---------------------------------------------------------------------------
# Page Config
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="ADMV Checker",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ---------------------------------------------------------------------------
# Custom CSS
# ---------------------------------------------------------------------------

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@300;400;500&display=swap');

/* ── Base ── */
html, body, [class*="css"] {
    font-family: 'DM Mono', monospace;
}

.stApp {
    background: #0a0a0f;
    color: #e8e6f0;
}

/* ── Hide default streamlit chrome ── */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 2rem 3rem 4rem 3rem; max-width: 1100px; }

/* ── Hero header ── */
.hero {
    text-align: center;
    padding: 3.5rem 0 2.5rem 0;
    position: relative;
}
.hero-badge {
    display: inline-block;
    background: rgba(99,102,241,0.15);
    border: 1px solid rgba(99,102,241,0.4);
    color: #a5b4fc;
    font-family: 'DM Mono', monospace;
    font-size: 0.7rem;
    letter-spacing: 0.2em;
    text-transform: uppercase;
    padding: 0.3rem 1rem;
    border-radius: 100px;
    margin-bottom: 1.2rem;
}
.hero-title {
    font-family: 'Syne', sans-serif;
    font-size: 3.8rem;
    font-weight: 800;
    line-height: 1.05;
    background: linear-gradient(135deg, #ffffff 0%, #a5b4fc 50%, #6366f1 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    margin: 0 0 0.8rem 0;
}
.hero-sub {
    color: #6b7280;
    font-size: 0.9rem;
    letter-spacing: 0.05em;
}

/* ── Upload zone ── */
.upload-zone {
    background: rgba(99,102,241,0.04);
    border: 1.5px dashed rgba(99,102,241,0.3);
    border-radius: 16px;
    padding: 2.5rem;
    margin: 2rem 0;
    transition: border-color 0.3s;
}
.upload-zone:hover { border-color: rgba(99,102,241,0.6); }

/* ── File info pill ── */
.file-pill {
    display: flex;
    align-items: center;
    gap: 0.8rem;
    background: rgba(16,185,129,0.08);
    border: 1px solid rgba(16,185,129,0.25);
    border-radius: 10px;
    padding: 0.9rem 1.2rem;
    margin: 1rem 0;
    font-size: 0.85rem;
    color: #6ee7b7;
}

/* ── Stat cards ── */
.stats-grid {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 0.8rem;
    margin: 1.5rem 0;
}
.stat-card {
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.07);
    border-radius: 12px;
    padding: 1.2rem 1rem;
    text-align: center;
}
.stat-card .stat-val {
    font-family: 'Syne', sans-serif;
    font-size: 2rem;
    font-weight: 700;
    color: #fff;
    line-height: 1;
}
.stat-card .stat-lbl {
    font-size: 0.65rem;
    color: #6b7280;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-top: 0.4rem;
}

/* ── ADMV breakdown cards ── */
.admv-grid {
    display: grid;
    grid-template-columns: repeat(5, 1fr);
    gap: 0.8rem;
    margin: 1rem 0 1.5rem 0;
}
.admv-card {
    border-radius: 12px;
    padding: 1.1rem;
    text-align: center;
    border: 1px solid transparent;
}
.admv-card.v { background: rgba(99,102,241,0.1);  border-color: rgba(99,102,241,0.3);  }
.admv-card.m { background: rgba(245,158,11,0.1);  border-color: rgba(245,158,11,0.3);  }
.admv-card.a { background: rgba(16,185,129,0.1);  border-color: rgba(16,185,129,0.3);  }
.admv-card.x { background: rgba(239,68,68,0.1);   border-color: rgba(239,68,68,0.3);   }
.admv-card.d { background: rgba(107,114,128,0.1); border-color: rgba(107,114,128,0.3); }
.admv-card .admv-letter {
    font-family: 'Syne', sans-serif;
    font-size: 1.8rem;
    font-weight: 800;
    line-height: 1;
}
.admv-card.v .admv-letter { color: #818cf8; }
.admv-card.m .admv-letter { color: #fbbf24; }
.admv-card.a .admv-letter { color: #34d399; }
.admv-card.x .admv-letter { color: #f87171; }
.admv-card.d .admv-letter { color: #9ca3af; }
.admv-card .admv-count {
    font-family: 'Syne', sans-serif;
    font-size: 1.4rem;
    font-weight: 700;
    color: #fff;
    margin-top: 0.3rem;
}
.admv-card .admv-lbl {
    font-size: 0.62rem;
    color: #6b7280;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-top: 0.2rem;
}

/* ── Section label ── */
.section-label {
    font-family: 'DM Mono', monospace;
    font-size: 0.65rem;
    color: #4b5563;
    text-transform: uppercase;
    letter-spacing: 0.2em;
    margin: 2rem 0 0.8rem 0;
    border-bottom: 1px solid rgba(255,255,255,0.05);
    padding-bottom: 0.5rem;
}

/* ── Log table ── */
.log-box {
    background: #0d0d14;
    border: 1px solid rgba(255,255,255,0.06);
    border-radius: 12px;
    padding: 1rem 1.2rem;
    max-height: 360px;
    overflow-y: auto;
    font-size: 0.78rem;
    line-height: 1.9;
}
.log-row { display: flex; gap: 1rem; align-items: baseline; }
.log-row .rn  { color: #374151; min-width: 60px; }
.log-row .tag { min-width: 90px; font-weight: 500; }
.log-row .tag.ok   { color: #34d399; }
.log-row .tag.skip { color: #f87171; }
.log-row .tag.leave{ color: #fbbf24; }
.log-row .st  { color: #6b7280; flex: 1; }
.log-row .ch  { color: #818cf8; min-width: 80px; text-align: right; }

/* ── Missing headers ── */
.missing-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr);
    gap: 0.5rem;
    margin: 1rem 0;
}
.missing-chip {
    background: rgba(239,68,68,0.08);
    border: 1px solid rgba(239,68,68,0.2);
    border-radius: 8px;
    padding: 0.5rem 0.8rem;
    font-size: 0.75rem;
    color: #fca5a5;
    font-family: 'DM Mono', monospace;
}

/* ── Buttons ── */
.stButton > button {
    background: linear-gradient(135deg, #6366f1, #4f46e5) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.7rem 2rem !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.95rem !important;
    letter-spacing: 0.03em !important;
    transition: all 0.2s !important;
    box-shadow: 0 4px 20px rgba(99,102,241,0.3) !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 28px rgba(99,102,241,0.45) !important;
}

/* ── Download button ── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #059669, #047857) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 600 !important;
    font-size: 1rem !important;
    padding: 0.8rem 2rem !important;
    box-shadow: 0 4px 20px rgba(5,150,105,0.3) !important;
    transition: all 0.2s !important;
}
.stDownloadButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 28px rgba(5,150,105,0.45) !important;
}

/* ── File uploader ── */
[data-testid="stFileUploader"] {
    background: transparent !important;
}
[data-testid="stFileUploaderDropzone"] {
    background: rgba(99,102,241,0.04) !important;
    border: 1.5px dashed rgba(99,102,241,0.3) !important;
    border-radius: 14px !important;
}

/* ── Spinner ── */
.stSpinner > div { border-top-color: #6366f1 !important; }

/* ── Alert boxes ── */
.stAlert { border-radius: 10px !important; }

/* ── Divider ── */
hr { border-color: rgba(255,255,255,0.05) !important; }

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: rgba(99,102,241,0.3); border-radius: 10px; }
</style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Hero
# ---------------------------------------------------------------------------

st.markdown("""
<div class="hero">
    <div class="hero-badge">⚡ Excel Automation Tool</div>
    <div class="hero-title">ADMV Checker</div>
    <div class="hero-sub">Upload · Validate · Process · Download</div>
</div>
""", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Upload
# ---------------------------------------------------------------------------

st.markdown('<div class="section-label">01 — Upload File</div>', unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Drop your Excel file here or click to browse",
    type=["xlsx", "xlsm"],
    label_visibility="visible"
)

if uploaded_file:
    original_name = os.path.splitext(uploaded_file.name)[0]
    output_name   = original_name + "_ADMV.xlsx"

    st.markdown(f"""
    <div class="file-pill">
        <span>📄</span>
        <span><strong>{uploaded_file.name}</strong> — ready to process</span>
        <span style="margin-left:auto; color:#4b5563;">→ output: <strong>{output_name}</strong></span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-label">02 — Run</div>', unsafe_allow_html=True)

    if st.button("⚡  Run ADMV Check", use_container_width=True):
        file_bytes = uploaded_file.read()

        with st.spinner("Validating headers and processing rows..."):
            success, output_bytes, logs, summary = process_excel(file_bytes)

        # ── Header validation failed ──
        if not success:
            st.markdown('<div class="section-label">⚠ Header Validation Failed</div>', unsafe_allow_html=True)
            st.error(f"**{len(summary['missing'])} missing header(s)** — processing stopped. No changes were made.")

            chips = "".join([f'<div class="missing-chip">{col}</div>' for col in summary["missing"]])
            st.markdown(f'<div class="missing-grid">{chips}</div>', unsafe_allow_html=True)

            if summary.get("extra"):
                with st.expander(f"ℹ️ {len(summary['extra'])} extra column(s) in file (not an error)"):
                    for col in summary["extra"]:
                        st.markdown(f"`{col}`")

        # ── Success ──
        else:
            st.markdown('<div class="section-label">03 — Results</div>', unsafe_allow_html=True)

            # Row summary
            s = summary
            st.markdown(f"""
            <div class="stats-grid">
                <div class="stat-card">
                    <div class="stat-val">{s['total_rows']}</div>
                    <div class="stat-lbl">Total Rows</div>
                </div>
                <div class="stat-card">
                    <div class="stat-val" style="color:#34d399">{s['processed']}</div>
                    <div class="stat-lbl">Processed</div>
                </div>
                <div class="stat-card">
                    <div class="stat-val" style="color:#f87171">{s['skipped']}</div>
                    <div class="stat-lbl">Skipped</div>
                </div>
                <div class="stat-card">
                    <div class="stat-val" style="color:#fbbf24">{s['left_as_is']}</div>
                    <div class="stat-lbl">Left As-Is</div>
                </div>
                <div class="stat-card">
                    <div class="stat-val" style="color:#818cf8">{s['changes']}</div>
                    <div class="stat-lbl">ADMV Updated</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            # ADMV breakdown
            st.markdown('<div class="section-label">ADMV Breakdown</div>', unsafe_allow_html=True)
            ac = s["admv_counts"]
            st.markdown(f"""
            <div class="admv-grid">
                <div class="admv-card v">
                    <div class="admv-letter">V</div>
                    <div class="admv-count">{ac['V']}</div>
                    <div class="admv-lbl">Verify</div>
                </div>
                <div class="admv-card m">
                    <div class="admv-letter">M</div>
                    <div class="admv-count">{ac['M']}</div>
                    <div class="admv-lbl">Modify</div>
                </div>
                <div class="admv-card a">
                    <div class="admv-letter">A</div>
                    <div class="admv-count">{ac['A']}</div>
                    <div class="admv-lbl">Add</div>
                </div>
                <div class="admv-card x">
                    <div class="admv-letter">X</div>
                    <div class="admv-count">{ac['X']}</div>
                    <div class="admv-lbl">Not Found</div>
                </div>
                <div class="admv-card d">
                    <div class="admv-letter">D</div>
                    <div class="admv-count">{ac['D']}</div>
                    <div class="admv-lbl">Delete</div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            if s["d_protected"] > 0:
                st.markdown(f"> 🔒 **{s['d_protected']}** ADMV cell(s) with existing `D` were protected and left untouched.")

            # Download
            st.markdown('<div class="section-label">04 — Download</div>', unsafe_allow_html=True)
            st.download_button(
                label=f"⬇️  Download  {output_name}",
                data=output_bytes,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            # Row logs
            st.markdown('<div class="section-label">05 — Row Log</div>', unsafe_allow_html=True)
            with st.expander("View detailed row-by-row log", expanded=False):
                rows_html = ""
                for kind, rnum, status, rchanges in logs:
                    tag_cls  = kind
                    tag_text = "PROCESSED" if kind == "ok" else ("SKIPPED" if kind == "skip" else "LEFT AS-IS")
                    ch_text  = f"+{rchanges} cells" if kind == "ok" else "—"
                    rows_html += f"""
                    <div class="log-row">
                        <span class="rn">Row {rnum}</span>
                        <span class="tag {tag_cls}">{tag_text}</span>
                        <span class="st">{status[:60] + '…' if len(status) > 60 else status}</span>
                        <span class="ch">{ch_text}</span>
                    </div>"""
                st.markdown(f'<div class="log-box">{rows_html}</div>', unsafe_allow_html=True)

else:
    # Empty state
    st.markdown("""
    <div style="text-align:center; padding: 3rem 0; color: #374151;">
        <div style="font-size:3rem; margin-bottom:1rem;">📂</div>
        <div style="font-family:'Syne',sans-serif; font-size:1rem; color:#4b5563;">
            Select an Excel file above to get started
        </div>
        <div style="font-size:0.75rem; color:#374151; margin-top:0.5rem;">
            Supports .xlsx and .xlsm formats
        </div>
    </div>
    """, unsafe_allow_html=True)

# Footer
st.markdown("""
<div style="text-align:center; margin-top:4rem; color:#1f2937; font-size:0.7rem; letter-spacing:0.1em;">
    ADMV CHECKER &nbsp;·&nbsp; SAFE CELL-LEVEL UPDATES &nbsp;·&nbsp; DATA IS NEVER MODIFIED
</div>
""", unsafe_allow_html=True)